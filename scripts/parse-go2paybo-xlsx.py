import hashlib
import json
import sys
from datetime import date, datetime, time
from pathlib import Path

from openpyxl import load_workbook


def text(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def iso_date(value):
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    raw = text(value)
    if not raw:
        return None
    for sep in ("/", "-", "."):
        parts = raw.split(sep)
        if len(parts) == 3:
            try:
                a, b, c = [int(part) for part in parts]
                if a > 2400:
                    a -= 543
                    return f"{a:04d}-{b:02d}-{c:02d}"
                if c > 2400:
                    c -= 543
                    return f"{c:04d}-{b:02d}-{a:02d}"
                if a > 31:
                    return f"{a:04d}-{b:02d}-{c:02d}"
                return f"{c:04d}-{b:02d}-{a:02d}"
            except ValueError:
                pass
    return raw[:10]


def iso_time(value):
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.time().replace(microsecond=0).isoformat()
    if isinstance(value, time):
        return value.replace(microsecond=0).isoformat()
    raw = text(value)
    return raw or None


def number(value):
    if value in (None, ""):
        return 0
    if isinstance(value, (int, float)):
        return float(value)
    raw = text(value).replace(",", "").replace("฿", "")
    try:
        return float(raw)
    except ValueError:
        return 0


def import_key(sheet, row_no, payload):
    material = json.dumps(payload, ensure_ascii=False, sort_keys=True, default=str)
    digest = hashlib.sha1(material.encode("utf-8")).hexdigest()[:16]
    return f"xlsx:{sheet}:{row_no}:{digest}"


def non_empty(row):
    return any(cell not in (None, "") for cell in row)


def parse_sheet(ws, sheet_name, mapper):
    rows = []
    for row_no, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not non_empty(row):
            continue
        payload = mapper(row)
        if not payload:
            continue
        payload["import_key"] = import_key(sheet_name, row_no, payload)
        rows.append(payload)
    return rows


def parse_crypto(row):
    payload = {
        "date": iso_date(row[0]),
        "time": iso_time(row[1]),
        "source_account": text(row[2]) or None,
        "status": text(row[3]) or "ซื้อ USDT",
        "target_account": text(row[4]) or None,
        "amount_thb": round(number(row[5]), 2),
        "exchange_rate": round(number(row[6]), 6),
        "usdt": round(number(row[7]), 6),
        "note": text(row[8]) or None,
        "user_name": text(row[9]) or "admin",
        "slip_url": text(row[10]) or None,
    }
    if not payload["date"]:
        return None
    return payload


def parse_transfer(row):
    payload = {
        "date": iso_date(row[0]),
        "time": iso_time(row[1]),
        "source_account": text(row[2]) or None,
        "status": text(row[3]) or "โยกเงิน",
        "target_account": text(row[4]) or None,
        "amount": round(number(row[5]), 2),
        "fee": round(number(row[6]), 2),
        "user_name": text(row[7]) or "admin",
        "note": text(row[8]) or None,
        "slip_url": text(row[9]) or None,
    }
    if not payload["date"]:
        return None
    return payload


def parse_bogo2pay(row):
    actual = round(number(row[4]), 2)
    fee = round(number(row[5]), 2)
    net = round(number(row[6]), 2) if row[6] not in (None, "") else round(actual - fee, 2)
    payload = {
        "date": iso_date(row[0]),
        "time": iso_time(row[1]),
        "item": text(row[2]) or "Go2Pay",
        "type": text(row[3]) or "ฝาก",
        "actual_amount": actual,
        "fee": fee,
        "net_amount": net,
        "note": text(row[7]) or None,
        "user_name": text(row[8]) or "admin",
    }
    if not payload["date"]:
        return None
    return payload


def parse_balance(row):
    payload = {
        "date": iso_date(row[0]),
        "time": iso_time(row[1]),
        "account_name": text(row[2]) or "ไม่ระบุ",
        "balance_type": text(row[3]) or "บัญชีฝาก",
        "amount": round(number(row[4]), 2),
        "user_name": text(row[5]) or "admin",
        "note": text(row[6]) or None,
    }
    if not payload["date"]:
        return None
    return payload


def main():
    if len(sys.argv) < 3:
        raise SystemExit("usage: parse-go2paybo-xlsx.py input.xlsx output.json")
    input_path = Path(sys.argv[1])
    output_path = Path(sys.argv[2])
    wb = load_workbook(input_path, data_only=True, read_only=False)
    result = {
        "crypto_transactions": parse_sheet(wb["คริปโต"], "คริปโต", parse_crypto),
        "transfers": parse_sheet(wb["โยกเงิน"], "โยกเงิน", parse_transfer),
        "bogo2pay_transactions": parse_sheet(wb["BoGo2pay"], "BoGo2pay", parse_bogo2pay),
        "balances": parse_sheet(wb["ยอดคงเหลือ"], "ยอดคงเหลือ", parse_balance),
    }
    output_path.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({table: len(rows) for table, rows in result.items()}, ensure_ascii=False))


if __name__ == "__main__":
    main()
